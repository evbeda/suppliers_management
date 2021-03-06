from datetime import timedelta
from datetime import datetime

from openpyxl import Workbook
from tempfile import NamedTemporaryFile
from django.core.files.base import ContentFile
from django.http import (
    HttpResponse,
)
from django.utils.translation import gettext_lazy as _

import attr

from utils import COMMENT_TAXPAYER, COMMENT_TAXPAYER_INT_VALUES


class ExcelReportInputParams:
    model = attr.ib
    headers_attrs = attr.ib
    tab_name = attr.ib

    def __init__(self, model, tab_name, headers_attrs):
        self.headers_attrs = headers_attrs
        self.model = model
        self.tab_name = tab_name


def generate_xls(params):
    queryset = params.model
    headers = params.headers_attrs.keys()
    attrs = params.headers_attrs.values()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = params.tab_name
    row_num = 1
    for col_num, column_title in enumerate(headers, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for data in queryset:
        row_num += 1
        row = []

        for attribute in attrs:
            model_atribute = getattr(data, attribute)
            if callable(model_atribute):
                row.append(model_atribute())
            else:
                row.append(model_atribute)

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    with NamedTemporaryFile() as tmp:
        workbook.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
    return stream


def generate_response_xls(xls_file, file_name):
    response = HttpResponse(ContentFile(xls_file), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={date}-{file_name}.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
        file_name=file_name
    )
    return response


def get_field_changes(form, except_field, model_to_compare):
    result = ""
    if hasattr(form, 'cleaned_data'):
        form_data = form.cleaned_data
    else:
        form_data = form.data
        except_field.append('csrfmiddlewaretoken')
    for field in form_data:
        if form_data[field] is None:
            form_data[field] = ''
        if field not in except_field and str(form_data[field]) != str(model_to_compare.__dict__[field]):
            if 'file' in field or type(form_data[field]).__name__ == 'int' or type(
                    model_to_compare.__dict__[field]).__name__ == 'int':
                result = _(COMMENT_TAXPAYER_INT_VALUES).format(result, str(form.fields[field].label))
            else:
                result = _(COMMENT_TAXPAYER).format(result, str(form.fields[field].label),
                                                    model_to_compare.__dict__[field],
                                                    form_data[field])
    return result
